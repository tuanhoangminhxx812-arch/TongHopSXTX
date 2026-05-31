"""
Ứng dụng Tổng hợp Báo cáo INV hàng tháng
Công ty Điện lực Vũng Tàu - TCKT_PCVT
"""

import streamlit as st
import pandas as pd
import os
import zipfile
import re
import xml.etree.ElementTree as ET
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CẤU HÌNH CỐ ĐỊNH
# ─────────────────────────────────────────────
ACCOUNTS = {
    "H": "24231",
    "I": "24232",
    "J": "33195",
    "K": "2428",
    "L": "154163",
    "M": "627961",
    "N": "6416211",
    "O": "6276211",
}
ACCOUNT_COLS = list(ACCOUNTS.keys())  # H, I, J, K, L, M, N, O
ACCOUNT_CODES = list(ACCOUNTS.values())

# Nhận diện file theo từ khóa trong tên
FILE_PATTERNS = {
    "17A": ["INV-017A", "INV017A", "017A"],
    "11A": ["INV-011A", "INV011A", "011A"],
    "BU6A": ["GL-BU-006A", "BU006A", "BU-006A", "006A"],
    "39": ["INV-039", "INV039", "039A"],
}


# ─────────────────────────────────────────────
# HÀM ĐỌC FILE EXCEL (không dùng openpyxl để tránh lỗi)
# ─────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def get_shared_strings(_zip_obj):
    try:
        with _zip_obj.open("xl/sharedStrings.xml") as f:
            tree = ET.parse(f)
        root = tree.getroot()
        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        strings = []
        for si in root.findall("main:si", ns):
            texts = si.findall(".//main:t", ns)
            value = "".join(t.text or "" for t in texts)
            strings.append(value)
        return strings
    except Exception:
        return []


def read_sheet_raw(xlsx_path: str, sheet_name: str) -> dict:
    """Đọc 1 sheet từ file xlsx, trả về dict {row_num: {col_letter: value}}"""
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            shared = get_shared_strings(z)

            # Đọc workbook để lấy sheet ID
            with z.open("xl/workbook.xml") as f:
                wb_root = ET.fromstring(f.read().decode("utf-8"))

            # Relationships
            try:
                with z.open("xl/_rels/workbook.xml.rels") as f:
                    rels_root = ET.fromstring(f.read().decode("utf-8"))
                rels_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
                rels = {r.get("Id"): r.get("Target") for r in rels_root.findall("r:Relationship", rels_ns)}
            except Exception:
                rels = {}

            ns_wb = {
                "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
                "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            }

            sheet_rid = None
            for sheet in wb_root.findall(".//main:sheet", ns_wb):
                name = sheet.get("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}name") or sheet.get("name")
                if name == sheet_name:
                    sheet_rid = sheet.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                    break

            if not sheet_rid:
                return {}

            target = rels.get(sheet_rid, "")
            if not target.startswith("xl/"):
                target = "xl/" + target

            with z.open(target) as f:
                ws_root = ET.fromstring(f.read().decode("utf-8"))

            ns_ws = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            rows_data = {}
            for row in ws_root.findall(".//main:row", ns_ws):
                row_num = int(row.get("r", 0))
                rows_data[row_num] = {}
                for cell in row.findall("main:c", ns_ws):
                    ref = cell.get("r", "")
                    ctype = cell.get("t", "n")
                    v_elem = cell.find("main:v", ns_ws)
                    value = ""
                    if v_elem is not None and v_elem.text:
                        if ctype == "s":
                            try:
                                value = shared[int(v_elem.text)]
                            except Exception:
                                value = v_elem.text
                        else:
                            value = v_elem.text
                    col_match = re.match(r"([A-Z]+)", ref)
                    if col_match:
                        rows_data[row_num][col_match.group(1)] = value
            return rows_data
    except Exception as e:
        st.error(f"Lỗi đọc file {os.path.basename(xlsx_path)}: {e}")
        return {}


def get_sheet_names(xlsx_path: str) -> list:
    """Lấy danh sách tên sheets"""
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            with z.open("xl/workbook.xml") as f:
                wb_root = ET.fromstring(f.read().decode("utf-8"))
        ns_wb = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        names = []
        for sheet in wb_root.findall(".//main:sheet", ns_wb):
            n = sheet.get("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}name") or sheet.get("name")
            if n:
                names.append(n)
        return names
    except Exception:
        return []


def to_float(v) -> float:
    """Chuyển giá trị ô thành số thực"""
    if v is None or v == "":
        return 0.0
    try:
        s = str(v).replace(",", "").replace(" ", "").replace("\xa0", "")
        return float(s)
    except Exception:
        return 0.0


# ─────────────────────────────────────────────
# NHẬN DIỆN FILE
# ─────────────────────────────────────────────
def detect_file_type(filename: str) -> str | None:
    fn_upper = filename.upper()
    for ftype, patterns in FILE_PATTERNS.items():
        for p in patterns:
            if p.upper() in fn_upper:
                return ftype
    return None


# ─────────────────────────────────────────────
# XỬ LÝ FILE 17A → DataFrame cột AC, AE, AF, AG
# ─────────────────────────────────────────────
def process_17A(xlsx_path: str) -> pd.DataFrame:
    """
    Đọc sheet 17A từ file Báo cáo hoặc INV-017A nguồn.
    Trả về DataFrame: [ma_ho_so, tai_khoan, nhap, xuat, date, desc]
    """
    sheets = get_sheet_names(xlsx_path)
    # Tìm sheet chứa dữ liệu 17A
    sheet_name = None
    for s in sheets:
        su = s.upper()
        if "17A" in su or "INV_017" in su or "INV-017" in su or "017" in su:
            sheet_name = s
            break
    if not sheet_name and sheets:
        sheet_name = sheets[0]
    if not sheet_name:
        return pd.DataFrame()

    raw = read_sheet_raw(xlsx_path, sheet_name)
    records = []
    for row_num in sorted(raw.keys()):
        row = raw[row_num]
        desc = row.get("D", "").strip()
        col_ac = row.get("AC", "").strip()
        
        # Trích xuất mã hồ sơ bằng regex để hỗ trợ cả file nguồn gốc và file mẫu
        ma_ho_so = ""
        m_ac = re.search(r'\b(VTA\d{5}[A-Z]\d{5})\b', col_ac.upper())
        m_d = re.search(r'\b(VTA\d{5}[A-Z]\d{5})\b', desc.upper())
        
        if m_ac:
            ma_ho_so = m_ac.group(1)
        elif m_d:
            ma_ho_so = m_d.group(1)
            
        if not ma_ho_so:
            continue
            
        # AE trong master, E trong raw
        tai_khoan = row.get("AE", "").strip() or row.get("E", "").strip()
        
        # AF trong master, I trong raw
        nhap = to_float(row.get("AF", 0)) or to_float(row.get("I", 0))
        
        # AG trong master, L trong raw
        xuat = to_float(row.get("AG", 0)) or to_float(row.get("L", 0))
        
        # M trong master, A trong raw
        date = row.get("M", "").strip() or row.get("A", "").strip()
        
        records.append({
            "ma_ho_so": ma_ho_so,
            "tai_khoan": tai_khoan,
            "nhap": nhap,
            "xuat": xuat,
            "date": date,
            "desc": desc,
        })
    return pd.DataFrame(records)


# ─────────────────────────────────────────────
# XỬ LÝ FILE BU6A → DataFrame cột P, S, T
# ─────────────────────────────────────────────
def process_BU6A(xlsx_path: str) -> pd.DataFrame:
    """
    Đọc sheet BU6A từ file GL-BU-006A hoặc Báo cáo.
    Trả về DataFrame: [ma_ho_so, nhap_bu6a, xuat_bu6a, date, desc]
    """
    sheets = get_sheet_names(xlsx_path)
    sheet_name = None
    for s in sheets:
        su = s.upper()
        if "BU6A" in su or "006A" in su or "BU_006" in su or "GL" in su:
            sheet_name = s
            break
    if not sheet_name and sheets:
        sheet_name = sheets[0]
    if not sheet_name:
        return pd.DataFrame()

    raw = read_sheet_raw(xlsx_path, sheet_name)
    records = []
    for row_num in sorted(raw.keys()):
        row = raw[row_num]
        desc = row.get("D", "").strip()
        col_p = row.get("P", "").strip()
        
        ma_ho_so = ""
        m_p = re.search(r'\b(VTA\d{5}[A-Z]\d{5})\b', col_p.upper())
        m_d = re.search(r'\b(VTA\d{5}[A-Z]\d{5})\b', desc.upper())
        
        if m_p:
            ma_ho_so = m_p.group(1)
        elif m_d:
            ma_ho_so = m_d.group(1)
            
        if not ma_ho_so:
            continue
            
        # S trong master, F (Phát sinh Có) trong raw
        nhap = to_float(row.get("S", 0)) or to_float(row.get("F", 0))
        
        # T trong master, E (Phát sinh Nợ) trong raw
        xuat = to_float(row.get("T", 0)) or to_float(row.get("E", 0))
        
        # B trong cả master và raw
        date = row.get("B", "").strip()
        
        records.append({
            "ma_ho_so": ma_ho_so,
            "nhap_bu6a": nhap,
            "xuat_bu6a": xuat,
            "date": date,
            "desc": desc,
        })
    return pd.DataFrame(records)


# ─────────────────────────────────────────────
# XỬ LÝ FILE 11A → DataFrame cột A→J
# ─────────────────────────────────────────────
def process_11A(xlsx_path: str) -> pd.DataFrame:
    """
    Đọc sheet 11A.
    Trả về DataFrame các cột A→J dòng dữ liệu (bỏ header)
    """
    sheets = get_sheet_names(xlsx_path)
    sheet_name = None
    for s in sheets:
        su = s.upper()
        if "11A" in su or "011A" in su or "INV_011" in su:
            sheet_name = s
            break
    if not sheet_name and sheets:
        sheet_name = sheets[0]
    if not sheet_name:
        return pd.DataFrame()

    raw = read_sheet_raw(xlsx_path, sheet_name)
    records = []
    for row_num in sorted(raw.keys()):
        row = raw[row_num]
        # Lấy cột A→J
        rec = {chr(65 + i): row.get(chr(65 + i), "") for i in range(10)}
        # Chỉ lấy dòng có dữ liệu thực (cột A có nội dung)
        if any(v.strip() for v in rec.values()):
            rec["_row"] = row_num
            records.append(rec)
    return pd.DataFrame(records)


# ─────────────────────────────────────────────
# XỬ LÝ FILE 39 → DataFrame
# ─────────────────────────────────────────────
def process_39(xlsx_path: str) -> pd.DataFrame:
    """
    Đọc sheet 39.
    """
    sheets = get_sheet_names(xlsx_path)
    sheet_name = sheets[0] if sheets else None
    if not sheet_name:
        return pd.DataFrame()

    raw = read_sheet_raw(xlsx_path, sheet_name)
    records = []
    for row_num in sorted(raw.keys()):
        row = raw[row_num]
        col_h = row.get("H", "").strip()
        col_g = row.get("G", "").strip()
        
        ma_ho_so = ""
        m_h = re.search(r'\b(VTA\d{5}[A-Z]\d{5})\b', col_h.upper())
        m_g = re.search(r'\b(VTA\d{5}[A-Z]\d{5})\b', col_g.upper())
        
        if m_h:
            ma_ho_so = m_h.group(1)
        elif m_g:
            ma_ho_so = m_g.group(1)
            
        if not ma_ho_so:
            continue
            
        xuat = to_float(row.get("AA", 0))
        nhap = to_float(row.get("AC", 0))
        
        if xuat == 0 and nhap == 0:
            xuat = to_float(row.get("O", 0))
            
        records.append({
            "ma_ho_so": ma_ho_so,
            "aa_39": xuat,
            "ac_39": nhap,
        })
    return pd.DataFrame(records)


# ─────────────────────────────────────────────
# TẠO BẢNG TONGHOP
# ─────────────────────────────────────────────
def extract_pa_ttr_from_text(text: str) -> str:
    if not text:
        return ""
    text_upper = text.upper()
    match = re.search(r'(PA\s*\d+|TTR\s*\d+|TT\s*:\s*\d+|TT\s*\d+)', text_upper)
    if match:
        val = match.group(1)
        val = val.replace("TTR", "Ttr").replace("TT:", "TT").replace("TT", "Ttr")
        return val
    return ""


def build_tonghop(
    df_phan_loai: pd.DataFrame,
    df_17a: pd.DataFrame,
    df_bu6a: pd.DataFrame,
    df_39: pd.DataFrame,
) -> pd.DataFrame:
    """
    Kết hợp các DataFrame để tạo bảng tổng hợp.
    """
    rows = []

    # Map để tự động tra cứu ngày mở HS và PA/Ttr từ dữ liệu gốc
    mhs_dates = {}
    mhs_pa_ttr = {}

    if not df_17a.empty:
        for _, r in df_17a.iterrows():
            mhs = r["ma_ho_so"]
            d = r.get("date", "")
            desc = r.get("desc", "")
            if d:
                try:
                    d_val = float(d)
                    if mhs not in mhs_dates or d_val < float(mhs_dates[mhs]):
                        mhs_dates[mhs] = str(int(d_val))
                except:
                    if mhs not in mhs_dates:
                        mhs_dates[mhs] = d
            if desc and mhs not in mhs_pa_ttr:
                extracted = extract_pa_ttr_from_text(desc)
                if extracted:
                    mhs_pa_ttr[mhs] = extracted

    if not df_bu6a.empty:
        for _, r in df_bu6a.iterrows():
            mhs = r["ma_ho_so"]
            d = r.get("date", "")
            desc = r.get("desc", "")
            if d:
                try:
                    d_val = float(d)
                    if mhs not in mhs_dates or d_val < float(mhs_dates[mhs]):
                        mhs_dates[mhs] = str(int(d_val))
                except:
                    if mhs not in mhs_dates:
                        mhs_dates[mhs] = d
            if desc and mhs not in mhs_pa_ttr:
                extracted = extract_pa_ttr_from_text(desc)
                if extracted:
                    mhs_pa_ttr[mhs] = extracted

    # Index lookup cho df_17a: (ma_ho_so, tai_khoan) → (nhap, xuat)
    map_17a_tk = {}
    map_17a_total = {}  # ma_ho_so → (tong_nhap, tong_xuat)
    if not df_17a.empty:
        for _, r in df_17a.iterrows():
            mhs = r["ma_ho_so"]
            tk = str(r["tai_khoan"]).strip()
            nhap = r["nhap"]
            xuat = r["xuat"]
            key = (mhs, tk)
            if key not in map_17a_tk:
                map_17a_tk[key] = {"nhap": 0.0, "xuat": 0.0}
            map_17a_tk[key]["nhap"] += nhap
            map_17a_tk[key]["xuat"] += xuat

            if mhs not in map_17a_total:
                map_17a_total[mhs] = {"nhap": 0.0, "xuat": 0.0}
            map_17a_total[mhs]["nhap"] += nhap
            map_17a_total[mhs]["xuat"] += xuat

    # Index BU6A: ma_ho_so → (nhap, xuat)
    map_bu6a = {}
    if not df_bu6a.empty:
        for _, r in df_bu6a.iterrows():
            mhs = r["ma_ho_so"]
            if mhs not in map_bu6a:
                map_bu6a[mhs] = {"nhap": 0.0, "xuat": 0.0}
            map_bu6a[mhs]["nhap"] += r["nhap_bu6a"]
            map_bu6a[mhs]["xuat"] += r["xuat_bu6a"]

    # Index 39
    map_39 = {}
    if not df_39.empty:
        for _, r in df_39.iterrows():
            mhs = r["ma_ho_so"]
            if mhs not in map_39:
                map_39[mhs] = {"aa": 0.0, "ac": 0.0}
            map_39[mhs]["aa"] += r["aa_39"]
            map_39[mhs]["ac"] += r["ac_39"]

    TEAM_MAP = {
        '1': "ĐỘI QUẢN LÝ HỆ THỐNG ĐO ĐẾM",
        '2': "ĐỘI QUẢN LÝ LƯỚI ĐIỆN",
        '3': "ĐỘI VẬN HÀNH LƯỚI ĐIỆN",
        '6': "ĐỘI TỔNG HỢP CÔN ĐẢO",
    }

    USER_MAP = {
        'N': "DIEU",
        'Y': "DIEU",
        'S': "SA",
        'R': "SA",
    }

    COST_TYPE_MAP = {
        'N': "GMDK",
        'Y': "BTDK",
        'S': "SCTX",
        'R': "SUCO",
    }

    for _, pl in df_phan_loai.iterrows():
        mhs = str(pl.get("Mã hồ sơ", "")).strip()
        if not mhs:
            continue

        section = str(pl.get("Phân loại", "")).strip()
        date_mo = pl.get("Date mở HS", "") if "Date mở HS" in df_phan_loai.columns else ""
        date_end = pl.get("Date END", "") if "Date END" in df_phan_loai.columns else ""
        user = pl.get("User", "") if "User" in df_phan_loai.columns else ""
        loai_cp = pl.get("Loại CP", "") if "Loại CP" in df_phan_loai.columns else ""
        doi = pl.get("Đội", "") if "Đội" in df_phan_loai.columns else ""
        pa_ttr = pl.get("PA/Ttr", "") if "PA/Ttr" in df_phan_loai.columns else ""
        quyet_toan = to_float(pl.get("Quyết toán", 0))

        # Tự động trích xuất thông tin từ Mã hồ sơ nếu trống
        if len(mhs) >= 9:
            char_7 = mhs[7]
            char_8 = mhs[8]
            
            if not user:
                user = USER_MAP.get(char_8, "")
            if not doi:
                doi = TEAM_MAP.get(char_7, "")
            if not loai_cp:
                loai_cp = COST_TYPE_MAP.get(char_8, "")

        if not date_mo:
            date_mo = mhs_dates.get(mhs, "")
            
        if not pa_ttr:
            pa_ttr = mhs_pa_ttr.get(mhs, "")

        # Tính các cột H→O (SUMIFS theo tài khoản)
        col_values = {}
        for col_letter, tk_code in ACCOUNTS.items():
            key = (mhs, tk_code)
            if key in map_17a_tk:
                net = map_17a_tk[key]["xuat"] - map_17a_tk[key]["nhap"]
            else:
                net = 0.0
            col_values[col_letter] = net

        # P: Tổng H→O
        tong_hp = sum(col_values[c] for c in ACCOUNT_COLS)

        # Q: quyết toán (nhập tay)
        col_q = quyet_toan

        # R: P - Q
        col_r = tong_hp - col_q

        # T: INV017a chi phí = SUMIFS AF (nhập tài khoản M=627961)
        # Thực ra T = SUM cột L→O trong tonghop (các TK chi phí)
        # Theo công thức gốc: =SUM(L:O) trong row
        col_t = sum(col_values[c] for c in ["L", "M", "N", "O"])

        # U: BU6A chi phí
        bu6a_data = map_bu6a.get(mhs, {"nhap": 0.0, "xuat": 0.0})
        col_u = bu6a_data["xuat"] - bu6a_data["nhap"]

        # V: T - U
        col_v = col_t - col_u

        # W: INV017a tổng (tất cả TK)
        total_17a = map_17a_total.get(mhs, {"nhap": 0.0, "xuat": 0.0})
        col_w = total_17a["xuat"] - total_17a["nhap"]

        # X: P - W
        col_x = tong_hp - col_w

        # Y: INV039 = AA - AC
        data_39 = map_39.get(mhs, {"aa": 0.0, "ac": 0.0})
        col_y = data_39["aa"] - data_39["ac"]

        # Z: W - Y
        col_z = col_w - col_y

        row = {
            "Phân loại": section,
            "Date mở HS": date_mo,
            "Date END": date_end,
            "User": user,
            "Mã hồ sơ": mhs,
            "Loại CP": loai_cp,
            "Đội": doi,
            "PA/Ttr": pa_ttr,
        }

        # Cột H→O
        for col_letter, tk_code in ACCOUNTS.items():
            row[f"TK {tk_code}"] = col_values[col_letter]

        row["Tổng XN (P)"] = tong_hp
        row["Quyết toán (Q)"] = col_q
        row["INV-QToán (R)"] = col_r
        row["INV017a CP (T)"] = col_t
        row["BU6A CP (U)"] = col_u
        row["Chênh lệch (V)"] = col_v
        row["INV017a Tổng (W)"] = col_w
        row["Cl INV017a (X)"] = col_x
        row["INV039 (Y)"] = col_y
        row["Cl INV039 (Z)"] = col_z

        rows.append(row)

    return pd.DataFrame(rows)


# ─────────────────────────────────────────────
# XUẤT EXCEL ĐẸP
# ─────────────────────────────────────────────
def to_excel_bytes(df_tonghop: pd.DataFrame, thang_nam: str) -> bytes:
    output = BytesIO()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Màu sắc
    COLOR_HEADER_BG = "1F4E79"    # Xanh đậm header
    COLOR_HEADER_FG = "FFFFFF"    # Trắng
    COLOR_SECTION_BG = "2E75B6"   # Xanh section
    COLOR_SECTION_FG = "FFFFFF"
    COLOR_GIA_HAN = "E2EFDA"      # Xanh nhạt - Gia hạn
    COLOR_QUYET_TOAN = "FFF2CC"   # Vàng nhạt - Quyết toán
    COLOR_DA_QT = "FCE4D6"        # Cam nhạt - Đã quyết toán
    COLOR_TONG = "D6E4F0"         # Xanh nhạt - dòng tổng
    COLOR_ALT = "F5F5F5"          # Xám nhạt - xen kẽ

    thin = Side(style="thin", color="AAAAAA")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    medium = Side(style="medium", color="1F4E79")
    medium_border = Border(left=medium, right=medium, top=medium, bottom=medium)

    def make_header_style(bg=COLOR_HEADER_BG, fg=COLOR_HEADER_FG, bold=True, size=10, wrap=True):
        return {
            "font": Font(bold=bold, color=fg, size=size, name="Calibri"),
            "fill": PatternFill("solid", fgColor=bg),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=wrap),
            "border": Border(
                left=Side(style="thin", color="FFFFFF"),
                right=Side(style="thin", color="FFFFFF"),
                top=Side(style="thin", color="FFFFFF"),
                bottom=Side(style="medium", color="1F4E79"),
            ),
        }

    def apply_style(cell, style_dict):
        for k, v in style_dict.items():
            setattr(cell, k, v)

    # ── Sheet TONGHOP ──
    ws = wb.create_sheet("tonghop")
    ws.sheet_view.showGridLines = False

    # Tiêu đề
    ws.merge_cells("A1:Z1")
    ws["A1"] = f"TỔNG HỢP BÁO CÁO INV - {thang_nam.upper()}"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79", name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor="DEEAF1")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:Z2")
    ws["A2"] = f"Ngày tạo: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Công ty Điện lực Vũng Tàu"
    ws["A2"].font = Font(italic=True, size=9, color="7F7F7F", name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Dòng ghi chú tài khoản (dòng 3)
    ws["A3"] = "Tài khoản:"
    ws["A3"].font = Font(bold=True, size=9, color="1F4E79", name="Calibri")
    for i, (col_l, tk) in enumerate(ACCOUNTS.items()):
        cell = ws.cell(row=3, column=9 + i)  # Cột H = cột 8
        cell.value = tk
        cell.font = Font(bold=True, size=9, color="7030A0", name="Calibri")
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 14

    # Header cột (dòng 4)
    COLUMNS = [
        "Date mở HS", "Date END", "User", "Mã hồ sơ", "Loại CP", "Đội", "PA/Ttr",
        f"TK {ACCOUNT_CODES[0]}", f"TK {ACCOUNT_CODES[1]}", f"TK {ACCOUNT_CODES[2]}",
        f"TK {ACCOUNT_CODES[3]}", f"TK {ACCOUNT_CODES[4]}", f"TK {ACCOUNT_CODES[5]}",
        f"TK {ACCOUNT_CODES[6]}", f"TK {ACCOUNT_CODES[7]}",
        "Tổng XN\n(P)", "QT\n(Q)", "INV-QT\n(R)", "",
        "INV017a CP\n(T)", "BU6A CP\n(U)", "Cl\n(V)",
        "INV017a Tổng\n(W)", "Cl INV017a\n(X)", "INV039\n(Y)", "Cl INV039\n(Z)",
    ]

    header_style = make_header_style()
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_style(cell, header_style)

    ws.row_dimensions[4].height = 36

    # Độ rộng cột
    col_widths = [12, 12, 8, 18, 8, 28, 35,
                  13, 13, 13, 13, 13, 13, 13, 13,
                  13, 13, 13, 4,
                  13, 13, 13, 13, 13, 13, 13]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze
    ws.freeze_panes = "D5"

    # Dữ liệu
    current_row = 5
    sections_order = ["Gia hạn", "Quyết toán", "Đã quyết toán", ""]
    section_colors = {
        "Gia hạn": COLOR_GIA_HAN,
        "Quyết toán": COLOR_QUYET_TOAN,
        "Đã quyết toán": COLOR_DA_QT,
        "": COLOR_ALT,
    }

    for section in sections_order:
        df_sec = df_tonghop[df_tonghop["Phân loại"] == section].copy()
        if df_sec.empty:
            continue

        # Dòng tiêu đề section
        ws.merge_cells(f"A{current_row}:Z{current_row}")
        sec_label = section if section else "Không phân loại"
        cell = ws.cell(row=current_row, column=1, value=f"  ▶  {sec_label.upper()}")
        cell.font = Font(bold=True, size=11, color=COLOR_SECTION_FG, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=COLOR_SECTION_BG)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        section_start = current_row
        row_color = section_colors.get(section, COLOR_ALT)

        num_fmt = '#,##0'

        for row_idx, (_, data_row) in enumerate(df_sec.iterrows()):
            bg = row_color if row_idx % 2 == 0 else "FFFFFF"
            fill = PatternFill("solid", fgColor=bg)

            def write_cell(col, value, is_num=False, bold=False, color_fg="000000"):
                c = ws.cell(row=current_row, column=col, value=value)
                c.fill = fill
                c.border = thin_border
                c.font = Font(name="Calibri", size=10, bold=bold, color=color_fg)
                if is_num:
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    if value != "" and value is not None:
                        c.number_format = num_fmt
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
                return c

            # Chuyển date
            def fmt_date(v):
                try:
                    if v and str(v).strip():
                        d_num = float(str(v).strip())
                        # Excel date serial
                        from datetime import date, timedelta
                        base = date(1899, 12, 30)
                        return (base + timedelta(days=int(d_num))).strftime("%d/%m/%Y")
                except Exception:
                    pass
                return str(v) if v else ""

            write_cell(1, fmt_date(data_row.get("Date mở HS", "")))
            write_cell(2, fmt_date(data_row.get("Date END", "")))
            write_cell(3, data_row.get("User", ""))
            c_mhs = write_cell(4, data_row.get("Mã hồ sơ", ""), bold=True, color_fg="1F4E79")
            write_cell(5, data_row.get("Loại CP", ""))
            write_cell(6, data_row.get("Đội", ""))
            write_cell(7, data_row.get("PA/Ttr", ""))

            # Cột H→O (số liệu tài khoản)
            for i, tk_code in enumerate(ACCOUNT_CODES):
                val = data_row.get(f"TK {tk_code}", 0) or 0
                c = write_cell(8 + i, val if val != 0 else "", is_num=True)

            write_cell(16, data_row.get("Tổng XN (P)", 0) or 0, is_num=True, bold=True)
            write_cell(17, data_row.get("Quyết toán (Q)", 0) or 0, is_num=True)
            write_cell(18, data_row.get("INV-QToán (R)", 0) or 0, is_num=True)
            write_cell(19, "")  # cột trống
            write_cell(20, data_row.get("INV017a CP (T)", 0) or 0, is_num=True)
            write_cell(21, data_row.get("BU6A CP (U)", 0) or 0, is_num=True)
            cl_v = data_row.get("Chênh lệch (V)", 0) or 0
            c_v = write_cell(22, cl_v if cl_v != 0 else "", is_num=True, color_fg="C00000" if cl_v != 0 else "000000")
            write_cell(23, data_row.get("INV017a Tổng (W)", 0) or 0, is_num=True)
            cl_x = data_row.get("Cl INV017a (X)", 0) or 0
            write_cell(24, cl_x if cl_x != 0 else "", is_num=True, color_fg="C00000" if cl_x != 0 else "000000")
            write_cell(25, data_row.get("INV039 (Y)", 0) or 0, is_num=True)
            write_cell(26, data_row.get("Cl INV039 (Z)", 0) or 0, is_num=True)

            ws.row_dimensions[current_row].height = 16
            current_row += 1

        # Dòng tổng của section
        tong_fill = PatternFill("solid", fgColor=COLOR_TONG)
        ws.merge_cells(f"A{current_row}:G{current_row}")
        c_sum_label = ws.cell(row=current_row, column=1, value=f"  Tổng {sec_label}")
        c_sum_label.fill = tong_fill
        c_sum_label.font = Font(bold=True, size=10, color="1F4E79", name="Calibri")
        c_sum_label.alignment = Alignment(horizontal="left", vertical="center")
        c_sum_label.border = thin_border

        numeric_cols = [
            "Tổng XN (P)", "Quyết toán (Q)", "INV-QToán (R)",
            "INV017a CP (T)", "BU6A CP (U)", "Chênh lệch (V)",
            "INV017a Tổng (W)", "Cl INV017a (X)", "INV039 (Y)", "Cl INV039 (Z)",
        ] + [f"TK {tk}" for tk in ACCOUNT_CODES]

        col_map = {
            **{f"TK {ACCOUNT_CODES[i]}": 8 + i for i in range(8)},
            "Tổng XN (P)": 16, "Quyết toán (Q)": 17, "INV-QToán (R)": 18,
            "INV017a CP (T)": 20, "BU6A CP (U)": 21, "Chênh lệch (V)": 22,
            "INV017a Tổng (W)": 23, "Cl INV017a (X)": 24, "INV039 (Y)": 25, "Cl INV039 (Z)": 26,
        }

        for col_name, col_idx in col_map.items():
            total_val = df_sec[col_name].fillna(0).sum() if col_name in df_sec.columns else 0
            c = ws.cell(row=current_row, column=col_idx, value=total_val if total_val != 0 else "")
            c.fill = tong_fill
            c.font = Font(bold=True, size=10, color="1F4E79", name="Calibri")
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.number_format = num_fmt
            c.border = thin_border

        ws.row_dimensions[current_row].height = 18
        current_row += 2  # Khoảng cách

    # ── Các sheet tab nguồn ──
    for sec_name, color in [("Gia hạn", "2ECC71"), ("Quyết toán", "F39C12"), ("Đã quyết toán", "E74C3C")]:
        df_sec = df_tonghop[df_tonghop["Phân loại"] == sec_name]
        if df_sec.empty:
            continue

        ws_s = wb.create_sheet(sec_name[:15])
        ws_s.sheet_view.showGridLines = False

        ws_s.merge_cells("A1:Z1")
        ws_s["A1"] = f"{sec_name.upper()} - {thang_nam.upper()}"
        ws_s["A1"].font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
        ws_s["A1"].fill = PatternFill("solid", fgColor=color)
        ws_s["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_s.row_dimensions[1].height = 28

        hdr_style = make_header_style(bg=color)
        for col_idx, header in enumerate(COLUMNS, 1):
            c = ws_s.cell(row=2, column=col_idx, value=header)
            apply_style(c, hdr_style)
        ws_s.row_dimensions[2].height = 36

        for col_widths_i, w in enumerate(col_widths, 1):
            ws_s.column_dimensions[get_column_letter(col_widths_i)].width = w

        ws_s.freeze_panes = "D3"

        for row_idx, (_, data_row) in enumerate(df_sec.iterrows()):
            r = 3 + row_idx
            bg = "F9F9F9" if row_idx % 2 == 0 else "FFFFFF"
            fill = PatternFill("solid", fgColor=bg)
            # Viết dữ liệu tương tự sheet tonghop (rút gọn)
            vals = [
                fmt_date(data_row.get("Date mở HS", "")) if False else str(data_row.get("Date mở HS", "")) or "",
                str(data_row.get("Date END", "")) or "",
                data_row.get("User", ""),
                data_row.get("Mã hồ sơ", ""),
                data_row.get("Loại CP", ""),
                data_row.get("Đội", ""),
                data_row.get("PA/Ttr", ""),
            ] + [data_row.get(f"TK {tk}", 0) or 0 for tk in ACCOUNT_CODES] + [
                data_row.get("Tổng XN (P)", 0) or 0,
                data_row.get("Quyết toán (Q)", 0) or 0,
                data_row.get("INV-QToán (R)", 0) or 0,
                "",
                data_row.get("INV017a CP (T)", 0) or 0,
                data_row.get("BU6A CP (U)", 0) or 0,
                data_row.get("Chênh lệch (V)", 0) or 0,
                data_row.get("INV017a Tổng (W)", 0) or 0,
                data_row.get("Cl INV017a (X)", 0) or 0,
                data_row.get("INV039 (Y)", 0) or 0,
                data_row.get("Cl INV039 (Z)", 0) or 0,
            ]
            for col_i, v in enumerate(vals, 1):
                c = ws_s.cell(row=r, column=col_i, value=v)
                c.fill = fill
                c.border = thin_border
                c.font = Font(name="Calibri", size=10)
                if col_i > 7:
                    c.alignment = Alignment(horizontal="right")
                    if isinstance(v, (int, float)) and v != 0:
                        c.number_format = num_fmt
                else:
                    c.alignment = Alignment(horizontal="left")
            ws_s.row_dimensions[r].height = 15

    wb.save(output)
    return output.getvalue()


# ─────────────────────────────────────────────
# TẠO FILE PHAN LOAI MẪU
# ─────────────────────────────────────────────
def create_phan_loai_template(ho_so_data: list) -> bytes:
    """Tạo file phan_loai.xlsx mẫu"""
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "phan_loai"
    ws.sheet_view.showGridLines = False

    headers = [
        "Mã hồ sơ", "Date mở HS", "Date END", "User",
        "Loại CP", "Đội", "PA/Ttr", "Phân loại", "Quyết toán", "Ghi chú"
    ]
    SECTIONS = ["Gia hạn", "Quyết toán", "Đã quyết toán"]

    # Header
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="medium", color="FFFFFF"))
    ws.row_dimensions[1].height = 30

    # Độ rộng cột
    widths = [20, 14, 14, 10, 12, 35, 60, 18, 16, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Dữ liệu mẫu
    section_colors = {
        "Gia hạn": "E2EFDA",
        "Quyết toán": "FFF2CC",
        "Đã quyết toán": "FCE4D6",
        "": "F5F5F5",
    }

    thin = Side(style="thin", color="AAAAAA")
    thin_b = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_num = 2
    for rec in ho_so_data:
        section = rec.get("section", "")
        # Bỏ qua dòng phụ (mã hồ sơ trùng liền kề)
        if not rec.get("ma_ho_so", "").startswith("VTA"):
            continue
        # Lấy dòng đầu của mỗi mã hồ sơ (có Date mở HS)
        if not rec.get("date_mo"):
            continue

        bg = section_colors.get(section, "F5F5F5")
        fill = PatternFill("solid", fgColor=bg)

        def fmt_date_from_serial(v):
            try:
                if v and str(v).strip():
                    d_num = float(str(v).strip())
                    from datetime import date, timedelta
                    base = date(1899, 12, 30)
                    return (base + timedelta(days=int(d_num))).strftime("%d/%m/%Y")
            except Exception:
                pass
            return str(v) if v else ""

        values = [
            rec.get("ma_ho_so", ""),
            fmt_date_from_serial(rec.get("date_mo", "")),
            fmt_date_from_serial(rec.get("date_end", "")),
            rec.get("user", ""),
            rec.get("loai_cp", ""),
            rec.get("doi", ""),
            rec.get("pa_ttr", ""),
            section,
            rec.get("quyet_toan", ""),
            rec.get("ghi_chu", ""),
        ]

        for col, v in enumerate(values, 1):
            c = ws.cell(row=row_num, column=col, value=v)
            c.fill = fill
            c.font = Font(name="Calibri", size=10)
            c.border = thin_b
            if col == 8:  # Phân loại
                c.font = Font(name="Calibri", size=10, bold=True, color="1F4E79")
            if col == 9:  # Quyết toán
                c.alignment = Alignment(horizontal="right")
                if v and str(v).strip() and str(v).strip() != "0":
                    try:
                        c.value = float(str(v).replace(",", ""))
                        c.number_format = "#,##0"
                    except Exception:
                        pass
        ws.row_dimensions[row_num].height = 16
        row_num += 1

    # Data validation cho cột Phân loại
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1='"Gia hạn,Quyết toán,Đã quyết toán"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.sqref = f"H2:H{row_num + 100}"

    ws.freeze_panes = "A2"
    wb.save(output)
    return output.getvalue()


# ─────────────────────────────────────────────
# GIAO DIỆN STREAMLIT
# ─────────────────────────────────────────────
def main():
    st.set_page_config(
        page_title="Tổng hợp INV | Điện lực Vũng Tàu",
        page_icon="⚡",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # CSS tùy chỉnh
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', sans-serif;
    }

    .stApp {
        background: linear-gradient(135deg, #f0f4f8 0%, #e8f0fe 100%);
    }

    /* Force main body text color to deep black (#111111) aggressively */
    .main .block-container, 
    .main .block-container p, 
    .main .block-container span, 
    .main .block-container div, 
    .main .block-container label, 
    .main .block-container li, 
    .main .block-container h2, 
    .main .block-container h3, 
    .main .block-container h4, 
    .main .block-container h5, 
    .main .block-container h6 {
        color: #111111 !important;
    }

    /* Force all inputs to have solid black text and white background */
    input, select, textarea {
        color: #111111 !important;
        background-color: white !important;
    }

    /* Style placeholders to be visible and clear dark gray */
    ::placeholder {
        color: #7f8c8d !important;
        opacity: 1 !important;
    }
    :-ms-input-placeholder { color: #7f8c8d !important; }
    ::-ms-input-placeholder { color: #7f8c8d !important; }

    /* Header */
    .main-header {
        background: linear-gradient(135deg, #1F4E79 0%, #2E75B6 60%, #3498DB 100%);
        padding: 2rem 2.5rem;
        border-radius: 16px;
        margin-bottom: 1.5rem;
        box-shadow: 0 8px 32px rgba(31, 78, 121, 0.3);
    }
    .main-header h1 {
        color: white !important;
        font-size: 1.8rem;
        font-weight: 700;
        margin: 0;
        text-shadow: 0 2px 4px rgba(0,0,0,0.2);
    }
    .main-header p {
        color: rgba(255,255,255,0.9) !important;
        margin: 0.3rem 0 0 0;
        font-size: 0.95rem;
    }

    /* Cards */
    .stat-card {
        background: white;
        border-radius: 12px;
        padding: 1.2rem 1.5rem;
        box-shadow: 0 2px 12px rgba(0,0,0,0.08);
        border-left: 4px solid #2E75B6;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    .stat-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(0,0,0,0.12);
    }
    .stat-card .number { font-size: 2rem; font-weight: 700; color: #111111 !important; }
    .stat-card .label {
        font-size: 0.82rem; color: #111111 !important;
        text-transform: uppercase; letter-spacing: 0.05em;
        margin-top: 0.2rem; font-weight: 600;
    }
    .stat-card div, .stat-card li, .stat-card ul, .stat-card b {
        color: #111111 !important; font-size: 0.93rem;
    }

    /* Section badges */
    .badge-gia-han {
        background: #C6EFCE; color: #111111 !important;
        padding: 4px 12px; border-radius: 20px;
        font-size: 0.85rem; font-weight: 700;
        display: inline-block; margin-bottom: 0.5rem;
    }
    .badge-quyet-toan {
        background: #D9E1F2; color: #1F4E79 !important;
        padding: 3px 10px; border-radius: 20px;
        font-size: 0.8rem; font-weight: 600;
    }
    .badge-da-qt {
        background: #FCE4D6; color: #833C0B !important;
        padding: 3px 10px; border-radius: 20px;
        font-size: 0.8rem; font-weight: 600;
    }

    /* File list */
    .file-item {
        background: white;
        border: 1px solid #e0e7ef;
        border-radius: 8px;
        padding: 0.6rem 1rem;
        margin-bottom: 0.5rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    .file-type-badge {
        background: #2E75B6;
        color: white !important;
        border-radius: 6px;
        padding: 2px 8px;
        font-size: 0.75rem;
        font-weight: 700;
    }

    /* Sidebar Styling - Modern Light Theme with high contrast */
    section[data-testid="stSidebar"] {
        background-color: #f0f4f8 !important;
        background-image: none !important;
        border-right: 1px solid #d0dbe5;
    }
    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3,
    section[data-testid="stSidebar"] h4,
    section[data-testid="stSidebar"] h5,
    section[data-testid="stSidebar"] h6,
    section[data-testid="stSidebar"] span,
    section[data-testid="stSidebar"] p,
    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] li,
    section[data-testid="stSidebar"] div,
    section[data-testid="stSidebar"] small {
        color: #111111 !important;
    }
    
    /* Make file uploader always have a clear, high-contrast light background */
    [data-testid="stFileUploader"] {
        background-color: #ffffff !important;
        border: 1px solid #d0dbe5 !important;
        border-radius: 10px !important;
        padding: 8px !important;
    }
    
    [data-testid="stFileUploaderDropzone"] {
        background-color: #f8f9fa !important;
        border: 2px dashed #b0c4de !important;
        border-radius: 8px !important;
    }

    /* Force all text inside uploader to be dark black for absolute contrast */
    [data-testid="stFileUploader"] * {
        color: #111111 !important;
    }

    /* Style the Browse files button to look premium and clear */
    [data-testid="stFileUploader"] button {
        background: linear-gradient(135deg, #1F4E79, #2E75B6) !important;
        border: none !important;
        border-radius: 6px !important;
        font-weight: 600 !important;
        padding: 0.4rem 1rem !important;
        box-shadow: 0 2px 6px rgba(31, 78, 121, 0.2) !important;
        transition: all 0.2s ease !important;
    }
    
    [data-testid="stFileUploader"] button:hover {
        background: linear-gradient(135deg, #2E75B6, #3498DB) !important;
        transform: translateY(-1px) !important;
    }

    [data-testid="stFileUploader"] button * {
        color: #ffffff !important;
    }

    /* Sidebar Action button */
    section[data-testid="stSidebar"] .stButton button {
        background: linear-gradient(135deg, #1F4E79, #2E75B6) !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        padding: 0.6rem 1.2rem !important;
        box-shadow: 0 4px 12px rgba(31, 78, 121, 0.2) !important;
        transition: all 0.2s ease !important;
        width: 100%;
    }
    section[data-testid="stSidebar"] .stButton button:hover {
        transform: translateY(-1px) !important;
        box-shadow: 0 6px 16px rgba(31, 78, 121, 0.3) !important;
    }
    section[data-testid="stSidebar"] .stButton button * {
        color: white !important;
    }

    /* Tabs styling - Force all tab labels to be high-contrast dark colors */
    .stTabs [data-baseweb="tab"] {
        font-weight: 600 !important;
        padding: 0.6rem 1.2rem !important;
    }
    .stTabs [data-baseweb="tab"], 
    .stTabs [data-baseweb="tab"] * {
        color: #333333 !important; /* Inactive tabs in bold dark gray */
        font-size: 1.05rem !important;
    }
    .stTabs [data-baseweb="tab"][aria-selected="true"], 
    .stTabs [data-baseweb="tab"][aria-selected="true"] * {
        color: #1F4E79 !important; /* Active tab in brand dark blue */
        font-weight: 700 !important;
    }

    /* Metric */
    [data-testid="stMetricValue"] {
        font-size: 1.5rem !important;
        color: #1F4E79 !important;
        font-weight: 700 !important;
    }
    [data-testid="stMetricLabel"] {
        color: #555555 !important;
    }

    /* Dataframe */
    .stDataFrame {
        border-radius: 10px;
        overflow: hidden;
        box-shadow: 0 2px 12px rgba(0,0,0,0.06);
    }

    /* Download button */
    .stDownloadButton button {
        background: linear-gradient(135deg, #1F4E79, #2E75B6) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        padding: 0.6rem 1.5rem !important;
        font-weight: 600 !important;
        font-size: 0.95rem !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 4px 15px rgba(31, 78, 121, 0.3) !important;
    }
    .stDownloadButton button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 25px rgba(31, 78, 121, 0.4) !important;
    }
    .stDownloadButton button * {
        color: white !important;
    }

    /* Success/warning box */
    .info-box {
        background: linear-gradient(135deg, #E8F5E9, #F1F8E9);
        border-left: 4px solid #4CAF50;
        border-radius: 8px;
        padding: 1rem 1.2rem;
        margin: 0.5rem 0;
        color: #111111 !important;
    }
    .info-box * {
        color: #111111 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── SESSION STATE INITIALIZATION ──
    if "folder_path" not in st.session_state:
        st.session_state.folder_path = ""
    if "thang_nam" not in st.session_state:
        st.session_state.thang_nam = f"Tháng {datetime.now().month:02d}/{datetime.now().year}"
    if "df_phan_loai" not in st.session_state:
        st.session_state.df_phan_loai = None
    if "phan_loai_file_name" not in st.session_state:
        st.session_state.phan_loai_file_name = ""
    if "df_result" not in st.session_state:
        st.session_state.df_result = None
    if "navigation_menu" not in st.session_state:
        st.session_state.navigation_menu = "📖 Hướng dẫn"

    # Header
    st.markdown("""
    <div class="main-header">
        <h1>⚡ Tổng hợp Báo cáo INV</h1>
        <p>Công ty Điện lực Vũng Tàu · TCKT_PCVT · Tự động tổng hợp từ các file nguồn</p>
    </div>
    """, unsafe_allow_html=True)

    # ── SIDEBAR MENU ──
    with st.sidebar:
        st.markdown("### ⚡ MENU CHỨC NĂNG")
        st.markdown("---")
        
        menu = st.radio(
            "Chọn chức năng:",
            options=[
                "📖 Hướng dẫn",
                "⚙️ Nhập dữ liệu & Cài đặt",
                "📂 Dữ liệu nguồn",
                "📊 Kết quả tổng hợp"
            ],
            key="navigation_menu",
            label_visibility="collapsed"
        )
        
        st.markdown("---")
        st.markdown("""
        <div style='font-size: 0.78rem; opacity: 0.8; text-align: center; color: #111111 !important;'>
        💡 Bố cục mới: Nhập liệu ở bên phải, Menu dọc ở bên trái.<br>Bấm các mục để chuyển đổi.
        </div>
        """, unsafe_allow_html=True)

    # ── MAIN AREA ROUTER ──

    # ── Tab 1: Hướng dẫn ──
    if menu == "📖 Hướng dẫn":
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("""
            <div class='stat-card'>
            <div class='label'>Bước 1</div>
            <div style='margin-top:0.5rem; font-size:0.95rem;'>
            📁 Đặt tất cả file Excel nguồn vào 1 thư mục:
            <ul style='margin-top:0.5rem;'>
            <li><b>INV-017A-Tháng XX.xlsx</b></li>
            <li><b>INV-011A-Tháng XX.xlsx</b></li>
            <li><b>GL-BU-006A-Tháng XX.xlsx</b></li>
            <li><i>INV-039A-... (nếu có)</i></li>
            </ul>
            </div>
            </div>
            """, unsafe_allow_html=True)

        with col2:
            st.markdown("""
            <div class='stat-card'>
            <div class='label'>Bước 2</div>
            <div style='margin-top:0.5rem; font-size:0.95rem;'>
            📋 Cập nhật file <b>phan_loai.xlsx</b>:
            <ul style='margin-top:0.5rem;'>
            <li>Điền <b>Phân loại</b>: Gia hạn / Quyết toán / Đã quyết toán</li>
            <li>Điền <b>Quyết toán</b>: số tiền quyết toán</li>
            <li>Thêm mã hồ sơ mới nếu có</li>
            </ul>
            </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        col3, col4 = st.columns(2)
        with col3:
            st.markdown("""
            <div class='stat-card'>
            <div class='label'>Bước 3</div>
            <div style='margin-top:0.5rem; font-size:0.95rem;'>
            ⚙️ Chọn tab <b>⚙️ Nhập dữ liệu & Cài đặt</b> ở bên trái,
            nhập thời gian, thư mục và upload file phân loại, rồi bấm <b>BẮT ĐẦU XỬ LÝ DỮ LIỆU</b>.
            </div>
            </div>
            """, unsafe_allow_html=True)

        with col4:
            st.markdown("""
            <div class='stat-card'>
            <div class='label'>Bước 4</div>
            <div style='margin-top:0.5rem; font-size:0.95rem;'>
            📊 Hệ thống tự động xử lý và chuyển sang tab <b>📊 Kết quả tổng hợp</b> để xem báo cáo và tải file Excel về.
            </div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("#### 📋 Tài khoản cố định (Cột H→O)")
        df_acc = pd.DataFrame([
            {"Cột": col, "Tài khoản": tk, "Mô tả": "XUAT - NHAP theo tài khoản kho"}
            for col, tk in ACCOUNTS.items()
        ])
        st.dataframe(df_acc, use_container_width=True, hide_index=True)

    # ── Tab 2: Dữ liệu nguồn ──
    elif menu == "📂 Dữ liệu nguồn":
        st.markdown("### 📂 Dữ liệu nguồn & Kiểm tra thư mục")
        st.markdown("---")
        
        # Tải file phan_loai mẫu
        st.markdown("#### 📥 Tải file phân loại mẫu (tháng 2/2026)")

        SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
        SAMPLE_DATA_PATH = os.path.join(SCRIPT_DIR, "phan_loai_mau.json")

        if os.path.exists(SAMPLE_DATA_PATH):
            import json
            with open(SAMPLE_DATA_PATH, "r", encoding="utf-8") as f:
                sample_data = json.load(f)
            phan_loai_bytes = create_phan_loai_template(sample_data)
            st.download_button(
                "⬇️ Tải file phan_loai.xlsx mẫu (Tháng 2/2026)",
                data=phan_loai_bytes,
                file_name="phan_loai.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        st.markdown("---")
        st.markdown("#### 🔍 Kiểm tra thư mục dữ liệu")

        folder_path = st.session_state.folder_path
        if folder_path and os.path.isdir(folder_path):
            files = [f for f in os.listdir(folder_path) if f.lower().endswith(".xlsx")]
            if files:
                st.success(f"✅ Tìm thấy {len(files)} file Excel trong: {folder_path}")
                for fname in files:
                    ftype = detect_file_type(fname)
                    badge = ftype or "?"
                    color = {"17A": "#2E75B6", "11A": "#27AE60", "BU6A": "#E67E22", "39": "#9B59B6"}.get(badge, "#95A5A6")
                    size_kb = os.path.getsize(os.path.join(folder_path, fname)) // 1024
                    st.markdown(f"""
                    <div class='file-item'>
                        <span class='file-type-badge' style='background:{color};'>{badge}</span>
                        <span style='flex:1; font-size:0.9rem; color:#111111 !important;'>{fname}</span>
                        <span style='color:#7F7F7F; font-size:0.8rem;'>{size_kb:,} KB</span>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.warning("⚠️ Không tìm thấy file Excel nào trong thư mục")
        elif folder_path:
            st.error(f"❌ Không tìm thấy thư mục: {folder_path}")
        else:
            st.info("👈 Hãy nhập đường dẫn thư mục trong tab **⚙️ Nhập dữ liệu & Cài đặt** để kiểm tra dữ liệu nguồn ở đây")

    # ── Tab 3: Cài đặt & Nhập liệu ──
    elif menu == "⚙️ Nhập dữ liệu & Cài đặt":
        st.markdown("### ⚙️ Cài đặt cấu hình & Upload dữ liệu")
        st.markdown("---")
        
        col_left, col_right = st.columns(2)
        with col_left:
            st.markdown("#### 📅 1. Cài đặt thời gian & Thư mục")
            st.session_state.thang_nam = st.text_input(
                "📅 Tháng/Năm báo cáo",
                value=st.session_state.thang_nam,
                help="Ví dụ: Tháng 02/2026"
            )
            
            st.session_state.folder_path = st.text_input(
                "📁 Đường dẫn thư mục dữ liệu gốc",
                value=st.session_state.folder_path,
                placeholder=r"z:\Cty ĐL VungTau\1. TCKT_PCVT\4. Công nghệ AI\KeToan(SCTX)\T2",
                help="Nhập đường dẫn đầy đủ đến thư mục chứa các file Excel thô"
            )
            
        with col_right:
            st.markdown("#### 📋 2. Upload file phân loại")
            phan_loai_file = st.file_uploader(
                "Upload file phan_loai.xlsx",
                type=["xlsx"],
                help="Upload file excel phân loại hồ sơ (Gia hạn / Quyết toán / Đã quyết toán)"
            )
            
            if phan_loai_file is not None:
                try:
                    df_p = pd.read_excel(phan_loai_file, sheet_name="phan_loai", dtype=str)
                    st.session_state.df_phan_loai = df_p.fillna("")
                    st.session_state.phan_loai_file_name = phan_loai_file.name
                    st.success(f"✅ Đã tải file phân loại: **{phan_loai_file.name}** ({len(df_p)} dòng)")
                except Exception as e:
                    st.error(f"❌ Lỗi đọc file phân loại: {e}")
            elif st.session_state.phan_loai_file_name:
                st.info(f"✅ Đang sử dụng file phân loại đã tải: **{st.session_state.phan_loai_file_name}**")
                
        st.markdown("<br>", unsafe_allow_html=True)
        process_btn = st.button("▶   BẮT ĐẦU XỬ LÝ DỮ LIỆU", use_container_width=True, type="primary")

        if process_btn:
            errors = []
            if not st.session_state.folder_path:
                errors.append("❌ Chưa nhập đường dẫn thư mục dữ liệu")
            elif not os.path.isdir(st.session_state.folder_path):
                errors.append(f"❌ Thư mục không tồn tại: {st.session_state.folder_path}")
            if st.session_state.df_phan_loai is None:
                errors.append("❌ Chưa upload file phân loại")

            if errors:
                for e in errors:
                    st.error(e)
                return

            folder_path = st.session_state.folder_path
            df_phan_loai = st.session_state.df_phan_loai

            # Quét thư mục
            files_in_folder = [f for f in os.listdir(folder_path) if f.lower().endswith(".xlsx")]

            df_17a = pd.DataFrame()
            df_bu6a = pd.DataFrame()
            df_11a = pd.DataFrame()
            df_39 = pd.DataFrame()

            progress = st.progress(0)
            status = st.empty()
            total = len(files_in_folder)

            for i, fname in enumerate(files_in_folder):
                fpath = os.path.join(folder_path, fname)
                ftype = detect_file_type(fname)
                status.text(f"🔄 Đang xử lý: {fname} ...")

                if ftype == "17A":
                    with st.spinner(f"Đang đọc {fname}..."):
                        df_17a = process_17A(fpath)
                        st.success(f"✅ 17A: {len(df_17a)} dòng dữ liệu")
                elif ftype == "BU6A":
                    with st.spinner(f"Đang đọc {fname}..."):
                        df_bu6a = process_BU6A(fpath)
                        st.success(f"✅ BU6A: {len(df_bu6a)} dòng dữ liệu")
                elif ftype == "11A":
                    with st.spinner(f"Đang đọc {fname}..."):
                        df_11a = process_11A(fpath)
                        st.success(f"✅ 11A: {len(df_11a)} dòng dữ liệu")
                elif ftype == "39":
                    with st.spinner(f"Đang đọc {fname}..."):
                        df_39 = process_39(fpath)
                        st.success(f"✅ 39: {len(df_39)} dòng dữ liệu")

                progress.progress((i + 1) / total)

            status.empty()
            progress.empty()

            if df_17a.empty:
                st.warning("⚠️ Không tìm thấy file INV-017A hoặc file trống")

            # Build tonghop
            with st.spinner("⚙️ Đang tổng hợp dữ liệu..."):
                df_result = build_tonghop(df_phan_loai, df_17a, df_bu6a, df_39)

            if df_result.empty:
                st.error("❌ Không có dữ liệu để tổng hợp. Kiểm tra lại file phân loại và thư mục dữ liệu.")
                return

            st.session_state.df_result = df_result
            st.success("🎉 Xử lý và tổng hợp dữ liệu thành công!")
            
            # Show success card with transition button
            st.markdown("""
            <div style='background: linear-gradient(135deg, #E8F5E9, #C8E6C9); border-left: 5px solid #2E7D32; border-radius: 12px; padding: 1.5rem; margin-top: 1rem;'>
                <h4 style='color: #1B5E20 !important; margin: 0 0 0.5rem 0;'>🎉 Hoàn thành tổng hợp dữ liệu!</h4>
                <p style='color: #2E7D32 !important; margin: 0 0 1rem 0; font-size: 0.95rem;'>
                    Hồ sơ của tháng đã được tổng hợp thành công từ các file nguồn. Anh hãy bấm nút bên dưới để chuyển sang màn hình xem báo cáo và tải file Excel nhé!
                </p>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button("➡️ XEM KẾT QUẢ TỔNG HỢP", use_container_width=True):
                st.session_state.navigation_menu = "📊 Kết quả tổng hợp"
                st.rerun()

    # ── Tab 4: Kết quả tổng hợp ──
    elif menu == "📊 Kết quả tổng hợp":
        df_result = st.session_state.df_result
        thang_nam = st.session_state.thang_nam

        if df_result is None or df_result.empty:
            st.info("⚠️ Chưa có kết quả tổng hợp. Vui lòng chọn tab **⚙️ Nhập dữ liệu & Cài đặt** ở menu bên trái, điền thông tin và bấm **BẮT ĐẦU XỬ LÝ DỮ LIỆU** trước nhé anh!")
            return

        # Thống kê tổng quan
        st.markdown("### 📊 Tổng quan")
        col1, col2, col3, col4 = st.columns(4)

        total_hs = len(df_result)
        tong_xn = df_result["Tổng XN (P)"].sum()
        tong_qt = df_result["Quyết toán (Q)"].sum()
        cl_inv = df_result["Chênh lệch (V)"].sum()

        with col1:
            st.metric("📁 Tổng hồ sơ", f"{total_hs}")
        with col2:
            st.metric("💰 Tổng XN (P)", f"{tong_xn:,.0f} đ")
        with col3:
            st.metric("✅ Tổng QT (Q)", f"{tong_qt:,.0f} đ")
        with col4:
            st.metric("⚖️ Chênh lệch (V)", f"{cl_inv:,.0f} đ",
                      delta_color="inverse" if cl_inv != 0 else "off")

        # Phân loại stats
        st.markdown("### 📋 Theo phân loại")
        col_a, col_b, col_c = st.columns(3)
        for col, sec, badge_class in [
            (col_a, "Gia hạn", "badge-gia-han"),
            (col_b, "Quyết toán", "badge-quyet-toan"),
            (col_c, "Đã quyết toán", "badge-da-qt"),
        ]:
            df_s = df_result[df_result["Phân loại"] == sec]
            with col:
                st.markdown(f"<span class='{badge_class}'>{sec}</span>", unsafe_allow_html=True)
                st.metric(f"Số hồ sơ", len(df_s))
                st.metric(f"Tổng XN", f"{df_s['Tổng XN (P)'].sum():,.0f}")

        # Hiển thị bảng
        st.markdown("### 📄 Bảng tổng hợp")

        # Filter
        with st.expander("🔍 Bộ lọc", expanded=False):
            fcol1, fcol2, fcol3 = st.columns(3)
            with fcol1:
                filter_sec = st.multiselect(
                    "Phân loại",
                    options=["Gia hạn", "Quyết toán", "Đã quyết toán"],
                    default=["Gia hạn", "Quyết toán", "Đã quyết toán"],
                )
            with fcol2:
                all_users = sorted(df_result["User"].dropna().unique().tolist())
                filter_user = st.multiselect("User", options=all_users, default=all_users)
            with fcol3:
                all_loai = sorted(df_result["Loại CP"].dropna().unique().tolist())
                filter_loai = st.multiselect("Loại CP", options=all_loai, default=all_loai)

        df_display = df_result[
            df_result["Phân loại"].isin(filter_sec) &
            df_result["User"].isin(filter_user) &
            df_result["Loại CP"].isin(filter_loai)
        ].copy()

        st.dataframe(
            df_display.style.format({
                c: "{:,.0f}" for c in [
                    f"TK {tk}" for tk in ACCOUNT_CODES
                ] + ["Tổng XN (P)", "Quyết toán (Q)", "INV-QToán (R)",
                     "INV017a CP (T)", "BU6A CP (U)", "Chênh lệch (V)",
                     "INV017a Tổng (W)", "Cl INV017a (X)", "INV039 (Y)", "Cl INV039 (Z)"]
                if c in df_display.columns
            }).applymap(
                lambda v: "color: #C00000" if isinstance(v, (int, float)) and v != 0 and v == v else "",
                subset=["Chênh lệch (V)"] if "Chênh lệch (V)" in df_display.columns else []
            ),
            use_container_width=True,
            height=500,
        )

        # Xuất Excel
        st.markdown("### 📥 Xuất file Excel")

        with st.spinner("Đang tạo file Excel..."):
            excel_bytes = to_excel_bytes(df_result, thang_nam)

        filename = f"TongHop_INV_{thang_nam.replace('/', '_').replace(' ', '_')}.xlsx"
        st.download_button(
            label=f"⬇️ Tải xuống {filename}",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.markdown("""
        <div class='info-box'>
        ✅ File Excel bao gồm:<br>
        • Sheet <b>tonghop</b>: Toàn bộ dữ liệu tổng hợp với format đẹp<br>
        • Sheet <b>Gia hạn</b>, <b>Quyết toán</b>, <b>Đã quyết toán</b>: Dữ liệu từng phần
        </div>
        """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
